"""
╔══════════════════════════════════════════════════════════════════╗
║        CSE AUTO-UPDATER  —  Colombo Stock Exchange 🇱🇰          ║
║                                                                  ║
║  Run this every day:  python cse_updater.py                      ║
║                                                                  ║
║  What it does:                                                   ║
║   1. Fetches today's live prices from cse.lk                     ║
║   2. Saves each day's data into RAW DATA sheet automatically     ║
║   3. Analysis + Dashboard + Charts update instantly              ║
║   4. Never need to paste data manually again!                    ║
║                                                                  ║
║  Install once:  pip install requests openpyxl                    ║
╚══════════════════════════════════════════════════════════════════╝
"""

import requests
import json
import os
import sys
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────
# ✏️  SETTINGS — Edit these
# ─────────────────────────────────────────────────────────────────────

# The ONE stock whose history you want to track in RAW DATA / Analysis
PRIMARY_STOCK = "COMB.N0000"

# All stocks to show in the Live Prices sheet
STOCKS_TO_TRACK = [
    "COMB.N0000",   # Commercial Bank of Ceylon
    "LOLC.N0000",   # LOLC Holdings
    "DIAL.N0000",   # Dialog Axiata
    "JKH.N0000",    # John Keells Holdings
    "HNB.N0000",    # Hatton National Bank
    "SAMP.N0000",   # Sampath Bank
    "NTB.N0000",    # Nations Trust Bank
    "GRAN.N0000",   # Ceylon Grain Elevators
    "CTC.N0000",    # Ceylon Tobacco
    "CARS.N0000",   # Carsons Cumberbatch
]

EXCEL_FILE = "CSE_Stock_Trend_Analyzer.xlsx"
BASE_URL   = "https://www.cse.lk/api/"

# ─────────────────────────────────────────────────────────────────────
# COLOURS & HELPERS
# ─────────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
GREEN_TXT  = "375623"
GREEN_BG   = "C6EFCE"
RED_TXT    = "9C0006"
RED_BG     = "FFC7CE"
GOLD       = "FFD700"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY  = "595959"
ACCENT     = "70AD47"

def fill(h):    return PatternFill("solid", start_color=h, end_color=h)
def bdr():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, r, c, v, bg=DARK_BLUE, fg=WHITE, bold=True, size=11, align='center', wrap=False):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name='Arial', bold=bold, size=size, color=fg)
    cell.fill      = fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = bdr()
    return cell

def dcell(ws, r, c, v, bg=WHITE, fg="000000", bold=False, size=10,
          align='center', fmt=None, wrap=False):
    cell = ws.cell(r, c, v)
    cell.font      = Font(name='Arial', bold=bold, size=size, color=fg)
    cell.fill      = fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border    = bdr()
    if fmt: cell.number_format = fmt
    return cell

# ─────────────────────────────────────────────────────────────────────
# CSE API CALLS
# ─────────────────────────────────────────────────────────────────────
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36',
    'Referer':    'https://www.cse.lk/',
    'Origin':     'https://www.cse.lk',
    'Accept':     'application/json, text/plain, */*',
}

def api_post(endpoint, data=None):
    try:
        r = requests.post(BASE_URL + endpoint, data=data or {}, headers=HEADERS, timeout=12)
        if r.status_code == 200:
            return r.json()
    except Exception as e:
        print(f"   ⚠️  API error ({endpoint}): {e}")
    return None

def fetch_stock_info(symbol):
    data = api_post("companyInfoSummery", {"symbol": symbol})
    if data:
        info = data.get("reqSymbolInfo", {})
        return {
            "symbol":     info.get("symbol", symbol),
            "name":       info.get("name", "Unknown"),
            "last_price": info.get("lastTradedPrice", 0) or 0,
            "change":     info.get("change", 0) or 0,
            "change_pct": info.get("changePercentage", 0) or 0,
            "market_cap": info.get("marketCap", 0) or 0,
            "fetched_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "error":      None,
        }
    return {"symbol": symbol, "error": "Could not fetch"}

def fetch_top_gainers():
    data = api_post("topGainers")
    if not data: return []
    if isinstance(data, list): return data[:10]
    return data.get("reqTopGainers", [])[:10]

def fetch_top_losers():
    data = api_post("topLooses")
    if not data: return []
    if isinstance(data, list): return data[:10]
    return data.get("reqTopLooses", [])[:10]

def fetch_most_active():
    data = api_post("mostActiveTrades")
    if not data: return []
    if isinstance(data, list): return data[:10]
    return data.get("reqMostActiveTrades", [])[:10]

def fetch_trade_summary():
    """Returns today's OHLCV for all stocks as a dict keyed by symbol."""
    data = api_post("tradeSummary")
    trades = (data or {}).get("reqTradeSummery", [])
    result = {}
    for t in trades:
        sym = t.get("symbol","")
        if sym:
            result[sym] = t
    return result

# ─────────────────────────────────────────────────────────────────────
# RAW DATA — append today's row
# ─────────────────────────────────────────────────────────────────────

def get_raw_data_sheet(wb):
    """Return the RAW DATA worksheet."""
    name = "📊 RAW DATA"
    if name in wb.sheetnames:
        return wb[name]
    return None

def find_next_empty_row(ws):
    """Find the first empty row in the data area (starting from row 3)."""
    row = 3
    while ws.cell(row, 2).value is not None:
        row += 1
    return row

def today_already_logged(ws, today_str):
    """Check if today's date is already in the sheet."""
    row = 3
    while ws.cell(row, 2).value is not None:
        val = ws.cell(row, 2).value
        if str(val)[:10] == today_str:
            return True
        row += 1
    return False

def append_today_to_raw(wb, trade_summary):
    """
    Append today's OHLCV for PRIMARY_STOCK into RAW DATA sheet.
    This is called every day — it adds one new row each run.
    """
    ws = get_raw_data_sheet(wb)
    if ws is None:
        print("   ⚠️  RAW DATA sheet not found!")
        return False

    today_str = date.today().strftime("%Y-%m-%d")

    # Don't add duplicate if already ran today
    if today_already_logged(ws, today_str):
        print(f"   ℹ️  Today ({today_str}) already in RAW DATA — skipping duplicate.")
        return False

    # Get trade data for primary stock
    stock_trade = trade_summary.get(PRIMARY_STOCK, {})

    # Try to get OHLCV values from trade summary
    # CSE API field names vary — we try multiple possibilities
    close_p = (stock_trade.get("closingPrice") or
               stock_trade.get("lastTradedPrice") or
               stock_trade.get("price") or 0)
    open_p  = (stock_trade.get("openingPrice") or
               stock_trade.get("open") or close_p)
    high_p  = (stock_trade.get("highPrice") or
               stock_trade.get("high") or close_p)
    low_p   = (stock_trade.get("lowPrice") or
               stock_trade.get("low") or close_p)
    volume  = (stock_trade.get("volume") or
               stock_trade.get("totalVolume") or
               stock_trade.get("totalTrades") or 0)

    # If tradeSummary didn't give full OHLCV, use companyInfo for close price
    if close_p == 0:
        info = fetch_stock_info(PRIMARY_STOCK)
        close_p = info.get("last_price", 0)
        open_p = high_p = low_p = close_p

    if close_p == 0:
        print(f"   ⚠️  Could not get price for {PRIMARY_STOCK} — RAW DATA not updated.")
        return False

    # Find next empty row and write
    next_row = find_next_empty_row(ws)
    bg = WHITE if next_row % 2 == 0 else LIGHT_GRAY

    dcell(ws, next_row, 2, today_str,        bg=bg, fmt='YYYY-MM-DD')
    dcell(ws, next_row, 3, float(open_p),    bg=bg, fmt='#,##0.00')
    dcell(ws, next_row, 4, float(high_p),    bg=bg, fmt='#,##0.00')
    dcell(ws, next_row, 5, float(low_p),     bg=bg, fmt='#,##0.00')
    dcell(ws, next_row, 6, float(close_p),   bg=bg, fmt='#,##0.00', bold=True)
    dcell(ws, next_row, 7, int(volume),      bg=bg, fmt='#,##0')

    print(f"   ✅ Added to RAW DATA → {today_str}  Close: LKR {close_p:,.2f}  Vol: {int(volume):,}")
    return True

# ─────────────────────────────────────────────────────────────────────
# LIVE PRICES SHEET
# ─────────────────────────────────────────────────────────────────────

def write_live_prices_sheet(wb, stock_data):
    name = "🇱🇰 CSE LIVE PRICES"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD

    widths = [3, 18, 32, 15, 13, 13, 20, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 42
    ws.merge_cells('B1:H1')
    c = ws.cell(1, 2, "🇱🇰   CSE LIVE PRICES  —  Colombo Stock Exchange")
    c.font = Font(name='Arial', bold=True, size=16, color=WHITE)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 22
    ws.merge_cells('B2:H2')
    c = ws.cell(2, 2, f"Last updated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}   |   Source: cse.lk   |   Currency: LKR")
    c.font = Font(name='Arial', size=10, color=WHITE, italic=True)
    c.fill = fill(MID_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[3].height = 26
    for ci, h in enumerate(["Symbol","Company Name","Last Price (LKR)","Change (LKR)","Change %","Market Cap (LKR)","Fetched At"], 2):
        hcell(ws, 3, ci, h, bg=MID_BLUE)

    for ri, stock in enumerate(stock_data, start=4):
        ws.row_dimensions[ri].height = 20
        bg = WHITE if ri % 2 == 0 else LIGHT_GRAY

        if stock.get("error"):
            dcell(ws, ri, 2, stock["symbol"], bg=bg)
            ws.merge_cells(f'C{ri}:H{ri}')
            dcell(ws, ri, 3, f"⚠️  {stock['error']}", bg="FFF2CC", fg="7F6000")
            continue

        chg     = stock.get("change", 0) or 0
        chg_pct = stock.get("change_pct", 0) or 0
        cbg     = GREEN_BG if chg >= 0 else RED_BG
        cfg     = GREEN_TXT if chg >= 0 else RED_TXT
        # Normalize percentage
        pct_val = chg_pct / 100 if abs(chg_pct) > 1 else chg_pct

        dcell(ws, ri, 2, stock["symbol"],           bg=bg,  fg=DARK_BLUE, bold=True)
        dcell(ws, ri, 3, stock["name"],              bg=bg,  align='left')
        dcell(ws, ri, 4, stock["last_price"],        bg=bg,  fmt='#,##0.00', bold=True)
        dcell(ws, ri, 5, chg,                        bg=cbg, fg=cfg, fmt='#,##0.00;(#,##0.00);"-"')
        dcell(ws, ri, 6, pct_val,                    bg=cbg, fg=cfg, fmt='0.00%')
        dcell(ws, ri, 7, stock["market_cap"],        bg=bg,  fmt='#,##0')
        dcell(ws, ri, 8, stock["fetched_at"],        bg=bg,  fg=DARK_GRAY)

    # Summary row
    sr = 4 + len(stock_data) + 1
    ws.row_dimensions[sr].height = 18
    ws.merge_cells(f'B{sr}:H{sr}')
    c = ws.cell(sr, 2, f"✅  {len(stock_data)} stocks tracked  •  Run script daily to keep prices updated  •  python cse_updater.py")
    c.font = Font(name='Arial', size=9, color=DARK_GRAY, italic=True)
    c.fill = fill(LIGHT_GRAY)
    c.alignment = Alignment(horizontal='center', vertical='center')

# ─────────────────────────────────────────────────────────────────────
# GAINERS & LOSERS SHEET
# ─────────────────────────────────────────────────────────────────────

def write_gainers_losers_sheet(wb, gainers, losers, active):
    name = "📊 GAINERS & LOSERS"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 1)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT

    for col, w in enumerate([3,16,28,13,13,3,16,28,13,13], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 38
    ws.merge_cells('B1:K1')
    c = ws.cell(1, 2, f"📊  TODAY'S MARKET MOVERS  —  {date.today().strftime('%A, %d %B %Y')}")
    c.font = Font(name='Arial', bold=True, size=14, color=WHITE)
    c.fill = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[2].height = 8

    ws.merge_cells('B3:E3')
    g = ws.cell(3, 2, "🟢  TOP GAINERS")
    g.font = Font(name='Arial', bold=True, size=12, color=WHITE)
    g.fill = fill(GREEN_TXT)
    g.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 26

    ws.merge_cells('G3:J3')
    l = ws.cell(3, 7, "🔴  TOP LOSERS")
    l.font = Font(name='Arial', bold=True, size=12, color=WHITE)
    l.fill = fill(RED_TXT)
    l.alignment = Alignment(horizontal='center', vertical='center')

    for ci, h in enumerate(["Symbol","Company","Price (LKR)","Change %"], 2):
        hcell(ws, 4, ci, h, bg=GREEN_TXT)
    for ci, h in enumerate(["Symbol","Company","Price (LKR)","Change %"], 7):
        hcell(ws, 4, ci, h, bg=RED_TXT)
    ws.row_dimensions[4].height = 22

    max_rows = max(len(gainers), len(losers), 1)
    for i in range(max_rows):
        ri = i + 5
        ws.row_dimensions[ri].height = 18
        bg_g = "E2EFDA" if i % 2 == 0 else GREEN_BG
        bg_l = "FFE7E7" if i % 2 == 0 else RED_BG

        if i < len(gainers):
            g = gainers[i]
            pct = g.get("changePercentage", g.get("percentageChange", 0)) or 0
            pct_v = pct/100 if abs(pct) > 1 else pct
            dcell(ws, ri, 2, g.get("symbol",""),         bg=bg_g, bold=True, fg=GREEN_TXT)
            dcell(ws, ri, 3, g.get("name",""),           bg=bg_g, align='left')
            dcell(ws, ri, 4, g.get("lastTradedPrice",0), bg=bg_g, fmt='#,##0.00')
            dcell(ws, ri, 5, pct_v,                      bg=bg_g, fg=GREEN_TXT, fmt='0.00%', bold=True)

        if i < len(losers):
            l = losers[i]
            pct = l.get("changePercentage", l.get("percentageChange", 0)) or 0
            pct_v = pct/100 if abs(pct) > 1 else pct
            dcell(ws, ri, 7,  l.get("symbol",""),         bg=bg_l, bold=True, fg=RED_TXT)
            dcell(ws, ri, 8,  l.get("name",""),           bg=bg_l, align='left')
            dcell(ws, ri, 9,  l.get("lastTradedPrice",0), bg=bg_l, fmt='#,##0.00')
            dcell(ws, ri, 10, pct_v,                      bg=bg_l, fg=RED_TXT, fmt='0.00%', bold=True)

    if active:
        start = max_rows + 7
        ws.merge_cells(f'B{start}:E{start}')
        a = ws.cell(start, 2, "🔥  MOST ACTIVE TRADES  (by Volume)")
        a.font = Font(name='Arial', bold=True, size=12, color=WHITE)
        a.fill = fill(MID_BLUE)
        a.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[start].height = 26

        for ci, h in enumerate(["Symbol","Company","Price (LKR)","Volume"], 2):
            hcell(ws, start+1, ci, h, bg=MID_BLUE)

        for i, s in enumerate(active):
            ri = start + 2 + i
            ws.row_dimensions[ri].height = 18
            bg = WHITE if i % 2 == 0 else LIGHT_GRAY
            dcell(ws, ri, 2, s.get("symbol",""),         bg=bg, bold=True, fg=DARK_BLUE)
            dcell(ws, ri, 3, s.get("name",""),           bg=bg, align='left')
            dcell(ws, ri, 4, s.get("lastTradedPrice",0), bg=bg, fmt='#,##0.00')
            dcell(ws, ri, 5, s.get("volume",0),          bg=bg, fmt='#,##0')

# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

def main():
    print("=" * 62)
    print("  🇱🇰  CSE AUTO-UPDATER  —  Colombo Stock Exchange")
    print(f"  📅  {datetime.now().strftime('%A, %d %B %Y  —  %H:%M:%S')}")
    print("=" * 62)

    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌  Excel file '{EXCEL_FILE}' not found!")
        print("    Make sure it is in the same folder as this script.")
        sys.exit(1)

    print(f"\n📂  Opening: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE)

    # ── 1. Fetch live prices ─────────────────────────────────
    print(f"\n📡  Fetching live prices for {len(STOCKS_TO_TRACK)} stocks...")
    stock_data = []
    for symbol in STOCKS_TO_TRACK:
        print(f"    → {symbol:<16}", end=" ")
        info = fetch_stock_info(symbol)
        if info.get("error"):
            print(f"❌  {info['error']}")
        else:
            arrow = "▲" if info['change'] >= 0 else "▼"
            print(f"✅  LKR {info['last_price']:>8,.2f}  {arrow} {info['change']:+.2f}  ({info['change_pct']:+.2f}%)")
        stock_data.append(info)

    # ── 2. Fetch trade summary (for RAW DATA OHLCV) ──────────
    print(f"\n📊  Fetching today's trade summary for RAW DATA...")
    trade_summary = fetch_trade_summary()
    if trade_summary:
        print(f"    → Got data for {len(trade_summary)} stocks")
    else:
        print("    → Trade summary unavailable — will use closing price only")

    # ── 3. Append today to RAW DATA ──────────────────────────
    print(f"\n📝  Updating RAW DATA sheet ({PRIMARY_STOCK})...")
    appended = append_today_to_raw(wb, trade_summary)

    # ── 4. Fetch market movers ───────────────────────────────
    print(f"\n📈  Fetching market movers...")
    gainers = fetch_top_gainers()
    losers  = fetch_top_losers()
    active  = fetch_most_active()
    print(f"    → Gainers: {len(gainers)}  Losers: {len(losers)}  Active: {len(active)}")

    # ── 5. Write sheets ──────────────────────────────────────
    print(f"\n💾  Writing sheets to Excel...")
    write_live_prices_sheet(wb, stock_data)
    print("    ✅  Live Prices sheet updated")
    write_gainers_losers_sheet(wb, gainers, losers, active)
    print("    ✅  Gainers & Losers sheet updated")

    # Set Live Prices as active sheet on open
    wb.active = wb["🇱🇰 CSE LIVE PRICES"]

    wb.save(EXCEL_FILE)

    # ── Summary ──────────────────────────────────────────────
    print("\n" + "=" * 62)
    print("  ✅  DONE!  Excel file saved successfully.")
    print(f"  📂  File: {os.path.abspath(EXCEL_FILE)}")

    # Count rows in RAW DATA
    if "📊 RAW DATA" in wb.sheetnames:
        ws = wb["📊 RAW DATA"]
        count = 0
        r = 3
        while ws.cell(r, 2).value:
            count += 1
            r += 1
        print(f"  📊  RAW DATA now has {count} days of price history")
        print(f"      Analysis & Dashboard auto-update from this data!")

    print("\n  💡  Run this script every trading day to keep building")
    print("      your price history automatically. The more days you")
    print("      run it, the better your charts become!")
    print("=" * 62)

if __name__ == "__main__":
    main()
